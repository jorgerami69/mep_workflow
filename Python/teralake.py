import os
import re
import pandas as pd
from docx import Document
import openpyxl

BASE_PATH = r"C:\data\workspace\mep_workflow\TeraLake"

print("📂 Iniciando escaneo en:", BASE_PATH)

total_files = 0
processed_files = 0
total_objects = 0

patterns = [
    r'\b([A-Z0-9_]+\.[A-Z0-9_]+)\b'
]

def extract_text_docx(path):
    try:
        doc = Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        print("❌ Error DOCX:", path, e)
        return ""

def extract_text_excel(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " ".join([str(cell) for cell in row if cell])
        return text
    except Exception as e:
        print("❌ Error Excel:", path, e)
        return ""

results = []

for root, dirs, files in os.walk(BASE_PATH):
    for file in files:
        total_files += 1
        path = os.path.join(root, file)

        text = ""
        tipo = ""

        if file.endswith(".docx"):
            text = extract_text_docx(path)
            tipo = "WORD"

        elif file.endswith(".xlsx") or file.endswith(".xlsm"):
            text = extract_text_excel(path)
            tipo = "EXCEL"

        else:
            continue

        processed_files += 1

        if text:
            matches = re.findall(patterns[0], text.upper())

            if matches:
                print(f"✅ Encontrado en {file}: {len(matches)} objetos")
                total_objects += len(matches)

                for obj in set(matches):
                    results.append({
                        "archivo": file,
                        "objeto": obj,
                        "tipo": tipo
                    })

print("\n📊 RESUMEN:")
print("Total archivos:", total_files)
print("Procesados:", processed_files)
print("Objetos encontrados:", total_objects)

df = pd.DataFrame(results)

if df.empty:
    print("⚠️ NO SE ENCONTRARON DATOS")
else:
    output_path = r"C:\data\workspace\mep_workflow\inventario_debug.csv"
    df.to_csv(output_path, index=False)
    print("✅ Archivo generado en:", output_path)