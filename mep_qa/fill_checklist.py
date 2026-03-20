#!/usr/bin/env python3
"""
Fill MEP_QA_Checklist.xlsx with QA results from findings.json.

Usage:
    python3 mep_qa/fill_checklist.py \
        --findings qa_reports/MEP_X/findings.json \
        --template MEP_QA_Checklist.xlsx \
        --output qa_reports/MEP_X/MEP_QA_Checklist_MEP_X.xlsx
"""

import argparse
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


# Map checklist item IDs to Excel rows
ITEM_ROWS = {
    "1.1": 7, "1.2": 8, "1.3": 9, "1.4": 10, "1.5": 11,
    "2.1": 13, "2.2": 14, "2.3": 15, "2.4": 16,
    "3.1": 18, "3.2": 19, "3.3": 20, "3.4": 21,
    "4.1": 23, "4.2": 24, "4.3": 25,
    "5.1": 27, "5.2": 28, "5.3": 29, "5.4": 30,
    "6.1": 32, "6.2": 33, "6.3": 34,
}


def fill_checklist(findings_path: str, template_path: str, output_path: str):
    """Fill the QA checklist Excel from findings JSON."""
    with open(findings_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Checklist QA"]

    mep = data.get("mep", {})
    score = data.get("score", {})
    checklist_auto = data.get("checklist_auto", {})
    checklist_observations = data.get("checklist_observations", {})
    qa_meta = data.get("qa_metadata", {})

    # --- Header ---
    ws["B2"] = mep.get("name", "")
    ws["D2"] = f"QA Automatizado (Claude + Scanner)"
    ws["F2"] = qa_meta.get("extraction_date", "")

    # SLA = reception + 1 business day
    reception = qa_meta.get("extraction_date", "")
    if reception:
        try:
            dt = datetime.strptime(reception, "%Y-%m-%d")
            sla = dt + timedelta(days=1)
            # Skip weekends
            while sla.weekday() >= 5:
                sla += timedelta(days=1)
            ws["B3"] = sla.strftime("%Y-%m-%d")
        except ValueError:
            pass

    # VALTX
    ws["D3"] = "No"  # Default; override if VALTX data present
    # F3 is formula — don't touch

    # --- Checklist items ---
    for item_id, row_num in ITEM_ROWS.items():
        status = checklist_auto.get(item_id)
        obs = checklist_observations.get(item_id, "")

        if status:
            ws.cell(row=row_num, column=4, value=status)
        if obs:
            ws.cell(row=row_num, column=5, value=obs)
        if status or obs:
            ws.cell(row=row_num, column=6, value="QA Auto")

    # --- MEP Tracker sheet ---
    if "MEP Tracker" in wb.sheetnames:
        ws2 = wb["MEP Tracker"]
        ws2["A3"] = mep.get("name", "")
        ws2["C3"] = qa_meta.get("extraction_date", "")
        ws2["E3"] = qa_meta.get("qa_date", "")
        ws2["F3"] = score.get("verdict", "")
        ws2["G3"] = score.get("verdict", "")
        ws2["H3"] = 0

        # Build summary for observations
        grade = score.get("grade", "?")
        overall = score.get("overall", "?")
        findings_list = data.get("findings", [])
        fails = [f for f in findings_list if f.get("severity") == "FAIL"]
        warns = [f for f in findings_list if f.get("severity") == "WARN"]
        summary_parts = [f"Score {overall}/100 ({grade})."]
        for f in fails[:3]:
            summary_parts.append(f["title"])
        ws2["I3"] = " | ".join(summary_parts)

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Checklist guardado en: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Fill MEP QA Checklist Excel from findings JSON"
    )
    parser.add_argument(
        "--findings", required=True,
        help="Path to findings.json from QA report",
    )
    parser.add_argument(
        "--template", required=True,
        help="Path to MEP_QA_Checklist.xlsx template",
    )
    parser.add_argument(
        "--output", required=True,
        help="Output path for filled checklist",
    )
    args = parser.parse_args()

    fill_checklist(args.findings, args.template, args.output)


if __name__ == "__main__":
    main()
